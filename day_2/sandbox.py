import os
import openpyxl

wb = openpyxl.load_workbook('day_2/lecture.xlsx')
print(str(wb))

new_sheet = wb.create_sheet()  #create new sheet

for i in range(1,5):
    wb.create_sheet()

print(str(wb.sheetnames))

print(new_sheet.columns)

# print("row index: " + str(wb.row))
# print("col index: " + str(wb.column))